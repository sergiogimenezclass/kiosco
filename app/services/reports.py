import sqlite3
import csv
import io
import datetime
from typing import Dict, Any, List, Tuple, Optional
from openpyxl import Workbook
from fpdf import FPDF

from app.core.errors import KioskException
from app.repositories.reports import ReportsRepository
from app.schemas.reports import VentasDiariasResponse, CajaReportItem, ProductoRankingItem
from app.schemas.catalog import ProductoResponse

class ReportsService:
    @staticmethod
    def get_ventas_diarias(
        db: sqlite3.Connection,
        desde: Optional[str] = None,
        hasta: Optional[str] = None
    ) -> Dict[str, Any]:
        return ReportsRepository.get_ventas_diarias_data(db, desde, hasta)

    @staticmethod
    def get_cajas(
        db: sqlite3.Connection,
        desde: Optional[str] = None,
        hasta: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        return ReportsRepository.get_cajas_data(db, desde, hasta)

    @staticmethod
    def get_ranking_productos(
        db: sqlite3.Connection,
        ordenar_por: str,
        limit: int = 15
    ) -> List[Dict[str, Any]]:
        if ordenar_por not in ["cantidad", "monto"]:
            raise KioskException(
                code="INVALID_PARAMETER",
                message="El parámetro 'ordenar_por' debe ser 'cantidad' o 'monto'",
                status_code=400
            )
        return ReportsRepository.get_ranking_productos(db, ordenar_por, limit)

    @staticmethod
    def get_stock_bajo(db: sqlite3.Connection) -> List[Dict[str, Any]]:
        return ReportsRepository.get_stock_bajo(db)

    @staticmethod
    def export_report(
        db: sqlite3.Connection,
        report_type: str,
        export_format: str,
        desde: Optional[str] = None,
        hasta: Optional[str] = None,
        ordenar_por: Optional[str] = "cantidad"
    ) -> Tuple[bytes, str, str]:
        """
        Generates and returns an export file.
        Returns: Tuple[content_bytes, filename, media_type]
        """
        export_format = export_format.lower()
        if export_format not in ["csv", "xlsx", "pdf"]:
            raise KioskException(
                code="INVALID_EXPORT_FORMAT",
                message="El formato de exportación debe ser 'csv', 'xlsx' o 'pdf'",
                status_code=400
            )

        if report_type == "ventas-diarias":
            data = ReportsRepository.get_ventas_diarias_data(db, desde, hasta)
            return ReportsService._generate_ventas_diarias_export(data, export_format, desde, hasta)
        
        elif report_type == "cajas":
            data = ReportsRepository.get_cajas_data(db, desde, hasta)
            return ReportsService._generate_cajas_export(data, export_format, desde, hasta)
            
        elif report_type == "ranking-productos":
            if ordenar_por not in ["cantidad", "monto"]:
                ordenar_por = "cantidad"
            data = ReportsRepository.get_ranking_productos(db, ordenar_por)
            return ReportsService._generate_ranking_export(data, export_format, ordenar_por)
            
        elif report_type == "stock-bajo":
            data = ReportsRepository.get_stock_bajo(db)
            return ReportsService._generate_stock_bajo_export(data, export_format)
            
        else:
            raise KioskException(
                code="INVALID_REPORT_TYPE",
                message=f"El tipo de reporte '{report_type}' no es válido",
                status_code=400
            )

    # --- GENERADORES DE ARCHIVOS ---

    @staticmethod
    def _generate_ventas_diarias_export(data: dict, fmt: str, desde: Optional[str], hasta: Optional[str]) -> Tuple[bytes, str, str]:
        filename = f"reporte_ventas_diarias_{desde or 'inicio'}_a_{hasta or 'fin'}"
        title = f"Reporte de Ventas Diarias ({desde or 'inicio'} a {hasta or 'fin'})"

        # Formateadores para centavos
        fmt_mon = lambda c: f"{c / 100:.2f}"

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow([title])
            writer.writerow([])
            writer.writerow(["RESUMEN GENERAL"])
            writer.writerow(["Metrica", "Valor"])
            writer.writerow(["Total General Facturado", fmt_mon(data["total_general_centavos"])])
            writer.writerow(["Cantidad de Ventas", data["cantidad_ventas"]])
            writer.writerow(["Descuentos Aplicados", fmt_mon(data["descuentos_applied_centavos"] if "descuentos_applied_centavos" in data else data.get("descuentos_aplicados_centavos", 0))])
            writer.writerow(["Cantidad de Anulaciones", data["cantidad_anulaciones"]])
            writer.writerow(["Total Anulado", fmt_mon(data["total_anulado_centavos"])])
            writer.writerow(["Cantidad de Devoluciones", data["cantidad_devoluciones"]])
            writer.writerow(["Total Devuelto", fmt_mon(data["total_devuelto_centavos"])])
            writer.writerow([])
            
            writer.writerow(["VENTAS POR METODO DE PAGO"])
            writer.writerow(["Metodo", "Total", "Cantidad"])
            for m in data["total_por_metodo"]:
                writer.writerow([m["metodo_pago"], fmt_mon(m["total_centavos"]), m["cantidad_ventas"]])
            writer.writerow([])
            
            writer.writerow(["VENTAS POR CAJERO"])
            writer.writerow(["Cajero", "Total", "Cantidad"])
            for c in data["total_por_cajero"]:
                writer.writerow([c["nombre_cajero"], fmt_mon(c["total_centavos"]), c["cantidad_ventas"]])
                
            return output.getvalue().encode("utf-8"), f"{filename}.csv", "text/csv"

        elif fmt == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas Diarias"
            
            ws.append([title])
            ws.append([])
            ws.append(["RESUMEN GENERAL"])
            ws.append(["Métrica", "Valor"])
            ws.append(["Total General Facturado", float(fmt_mon(data["total_general_centavos"]))])
            ws.append(["Cantidad de Ventas", data["cantidad_ventas"]])
            ws.append(["Descuentos Aplicados", float(fmt_mon(data.get("descuentos_aplicados_centavos", 0)))])
            ws.append(["Cantidad de Anulaciones", data["cantidad_anulaciones"]])
            ws.append(["Total Anulado", float(fmt_mon(data["total_anulado_centavos"]))])
            ws.append(["Cantidad de Devoluciones", data["cantidad_devoluciones"]])
            ws.append(["Total Devuelto", float(fmt_mon(data["total_devuelto_centavos"]))])
            ws.append([])
            
            ws.append(["VENTAS POR MÉTODO DE PAGO"])
            ws.append(["Método", "Total", "Cantidad"])
            for m in data["total_por_metodo"]:
                ws.append([m["metodo_pago"], float(fmt_mon(m["total_centavos"])), m["cantidad_ventas"]])
            ws.append([])
            
            ws.append(["VENTAS POR CAJERO"])
            ws.append(["Cajero", "Total", "Cantidad"])
            for c in data["total_por_cajero"]:
                ws.append([c["nombre_cajero"], float(fmt_mon(c["total_centavos"])), c["cantidad_ventas"]])
                
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue(), f"{filename}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else: # pdf
            pdf = PDFReport()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 14)
            pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align='L')
            pdf.ln(5)

            # Resumen
            pdf.set_font('helvetica', 'B', 12)
            pdf.cell(0, 8, "Resumen General", new_x="LMARGIN", new_y="NEXT", align='L')
            pdf.set_font('helvetica', '', 10)
            resumen_items = [
                ("Total General Facturado", f"${fmt_mon(data['total_general_centavos'])}"),
                ("Cantidad de Ventas", str(data["cantidad_ventas"])),
                ("Descuentos Aplicados", f"${fmt_mon(data.get('descuentos_aplicados_centavos', 0))}"),
                ("Cantidad de Anulaciones", str(data["cantidad_anulaciones"])),
                ("Total Anulado", f"${fmt_mon(data['total_anulado_centavos'])}"),
                ("Cantidad de Devoluciones", str(data["cantidad_devoluciones"])),
                ("Total Devuelto", f"${fmt_mon(data['total_devuelto_centavos'])}")
            ]
            for label, val in resumen_items:
                pdf.cell(70, 6, label, border=1)
                pdf.cell(50, 6, val, border=1, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)

            # Método de Pago
            pdf.set_font('helvetica', 'B', 12)
            pdf.cell(0, 8, "Ventas por Método de Pago", new_x="LMARGIN", new_y="NEXT", align='L')
            pdf.set_font('helvetica', 'B', 10)
            pdf.cell(50, 6, "Método", border=1)
            pdf.cell(40, 6, "Total", border=1)
            pdf.cell(30, 6, "Cantidad", border=1, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font('helvetica', '', 10)
            for m in data["total_por_metodo"]:
                pdf.cell(50, 6, m["metodo_pago"], border=1)
                pdf.cell(40, 6, f"${fmt_mon(m['total_centavos'])}", border=1)
                pdf.cell(30, 6, str(m["cantidad_ventas"]), border=1, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(5)

            # Por Cajero
            pdf.set_font('helvetica', 'B', 12)
            pdf.cell(0, 8, "Ventas por Cajero", new_x="LMARGIN", new_y="NEXT", align='L')
            pdf.set_font('helvetica', 'B', 10)
            pdf.cell(50, 6, "Cajero", border=1)
            pdf.cell(40, 6, "Total", border=1)
            pdf.cell(30, 6, "Cantidad", border=1, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font('helvetica', '', 10)
            for c in data["total_por_cajero"]:
                pdf.cell(50, 6, c["nombre_cajero"], border=1)
                pdf.cell(40, 6, f"${fmt_mon(c['total_centavos'])}", border=1)
                pdf.cell(30, 6, str(c["cantidad_ventas"]), border=1, new_x="LMARGIN", new_y="NEXT")

            return bytes(pdf.output()), f"{filename}.pdf", "application/pdf"

    @staticmethod
    def _generate_cajas_export(data: list, fmt: str, desde: Optional[str], hasta: Optional[str]) -> Tuple[bytes, str, str]:
        filename = f"reporte_cajas_{desde or 'inicio'}_a_{hasta or 'fin'}"
        title = f"Historial de Cajas ({desde or 'inicio'} a {hasta or 'fin'})"

        headers = [
            "ID", "Estado", "Apertura por", "Cierre por", "Monto Inicial", 
            "Ingresos", "Retiros", "Ventas Efectivo", "Ventas Digital", 
            "Declarado", "Esperado", "Desviación", "Fecha Apertura", "Fecha Cierre"
        ]

        fmt_mon = lambda c: f"{c / 100:.2f}" if c is not None else "0.00"
        fmt_mon_opt = lambda c: f"{c / 100:.2f}" if c is not None else ""
        fmt_fecha = lambda f: f.split(".")[0].replace("T", " ") if f else ""

        rows = []
        for c in data:
            rows.append([
                c["id"][:8],
                c["estado"],
                c["usuario_apertura_nombre"],
                c["usuario_cierre_nombre"] or "",
                fmt_mon(c["monto_inicial_centavos"]),
                fmt_mon(c["monto_ingresos_centavos"]),
                fmt_mon(c["monto_retiros_centavos"]),
                fmt_mon(c["monto_ventas_efectivo_centavos"]),
                fmt_mon(c["monto_ventas_digital_centavos"]),
                fmt_mon_opt(c["monto_declarado_centavos"]),
                fmt_mon_opt(c["monto_esperado_centavos"]),
                fmt_mon_opt(c["desviacion_centavos"]),
                fmt_fecha(c["fecha_apertura"]),
                fmt_fecha(c["fecha_cierre"])
            ])

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow([title])
            writer.writerow([])
            writer.writerow(headers)
            writer.writerows(rows)
            return output.getvalue().encode("utf-8"), f"{filename}.csv", "text/csv"

        elif fmt == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Cajas"
            ws.append([title])
            ws.append([])
            ws.append(headers)
            for r in rows:
                # Convertir montos a floats para formato numérico en Excel
                ws.append([
                    r[0], r[1], r[2], r[3],
                    float(r[4]), float(r[5]), float(r[6]), float(r[7]), float(r[8]),
                    float(r[9]) if r[9] else "", float(r[10]) if r[10] else "", float(r[11]) if r[11] else "",
                    r[12], r[13]
                ])
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue(), f"{filename}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else: # pdf
            pdf = PDFReport(orientation='L') # Horizontal para albergar muchas columnas
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 14)
            pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align='C')
            pdf.ln(5)

            # Como son 14 columnas, ajustemos anchos específicos para landscape (277mm disponibles)
            col_widths = [15, 18, 25, 25, 20, 15, 15, 23, 23, 20, 20, 20, 28, 28] # suma = 277
            
            pdf.set_font('helvetica', 'B', 8)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align='C')
            pdf.ln()

            pdf.set_font('helvetica', '', 7.5)
            for row in rows:
                for i, cell in enumerate(row):
                    val_str = str(cell)
                    if i in [4, 5, 6, 7, 8, 9, 10, 11] and val_str:
                        val_str = f"${val_str}"
                    pdf.cell(col_widths[i], 7, val_str, border=1, align='C')
                pdf.ln()

            return bytes(pdf.output()), f"{filename}.pdf", "application/pdf"

    @staticmethod
    def _generate_ranking_export(data: list, fmt: str, ordenar_por: str) -> Tuple[bytes, str, str]:
        filename = f"reporte_ranking_productos_por_{ordenar_por}"
        title = f"Ranking de Productos (Ordenado por {ordenar_por.capitalize()})"

        headers = ["Posición", "Producto ID", "Nombre Producto", "Cantidad Vendida", "Monto Vendido ($)"]
        
        fmt_mon = lambda c: f"{c / 100:.2f}"

        rows = []
        for i, item in enumerate(data, 1):
            rows.append([
                i,
                item["producto_id"][:8],
                item["nombre_producto"],
                item["cantidad_vendida"],
                fmt_mon(item["monto_sold_centavos"] if "monto_sold_centavos" in item else item.get("monto_vendido_centavos", 0))
            ])

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow([title])
            writer.writerow([])
            writer.writerow(headers)
            writer.writerows(rows)
            return output.getvalue().encode("utf-8"), f"{filename}.csv", "text/csv"

        elif fmt == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Ranking"
            ws.append([title])
            ws.append([])
            ws.append(headers)
            for r in rows:
                ws.append([r[0], r[1], r[2], r[3], float(r[4])])
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue(), f"{filename}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else: # pdf
            pdf = PDFReport()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 14)
            pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align='C')
            pdf.ln(5)

            col_widths = [20, 30, 80, 30, 30] # suma = 190
            
            pdf.set_font('helvetica', 'B', 10)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align='C')
            pdf.ln()

            pdf.set_font('helvetica', '', 9)
            for row in rows:
                for i, cell in enumerate(row):
                    val_str = str(cell)
                    if i == 4:
                        val_str = f"${val_str}"
                    pdf.cell(col_widths[i], 7, val_str, border=1, align='C')
                pdf.ln()

            return bytes(pdf.output()), f"{filename}.pdf", "application/pdf"

    @staticmethod
    def _generate_stock_bajo_export(data: list, fmt: str) -> Tuple[bytes, str, str]:
        filename = "reporte_stock_bajo"
        title = "Reporte de Productos con Stock Bajo Mínimo"

        headers = ["Producto ID", "Nombre Producto", "Stock Actual", "Stock Mínimo", "Categoría", "Unidad Medida"]

        rows = []
        for item in data:
            rows.append([
                item["producto_id"][:8],
                item["nombre_producto"],
                item["stock_actual"],
                item["stock_minimo"],
                item["categoria_nombre"],
                item["unidad_medida"]
            ])

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow([title])
            writer.writerow([])
            writer.writerow(headers)
            writer.writerows(rows)
            return output.getvalue().encode("utf-8"), f"{filename}.csv", "text/csv"

        elif fmt == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Bajo"
            ws.append([title])
            ws.append([])
            ws.append(headers)
            for r in rows:
                ws.append(r)
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue(), f"{filename}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else: # pdf
            pdf = PDFReport()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 14)
            pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align='C')
            pdf.ln(5)

            col_widths = [25, 75, 23, 23, 24, 20] # suma = 190
            
            pdf.set_font('helvetica', 'B', 9)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align='C')
            pdf.ln()

            pdf.set_font('helvetica', '', 9.5)
            for row in rows:
                for i, cell in enumerate(row):
                    pdf.cell(col_widths[i], 7, str(cell), border=1, align='C')
                pdf.ln()

            return bytes(pdf.output()), f"{filename}.pdf", "application/pdf"

# --- PDF Helper Class ---
class PDFReport(FPDF):
    def __init__(self, orientation='P'):
        super().__init__(orientation=orientation, unit='mm', format='A4')
        
    def header(self):
        self.set_font('helvetica', 'B', 10)
        self.cell(0, 8, 'KIOSCO POS - SISTEMA DE CONTROL DE VENTAS E INVENTARIO', border=0, align='L')
        self.set_font('helvetica', '', 9)
        self.cell(0, 8, datetime.datetime.now().strftime("%d/%m/%Y %H:%M"), border=0, align='R', new_x="LMARGIN", new_y="NEXT")
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')
